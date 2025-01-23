# app/main.py
from json import loads
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pathlib import Path
from typing import List, Dict
from pydantic import BaseModel
from fastapi.middleware.cors import CORSMiddleware

import pandas as pd
import tempfile
from openpyxl.utils import get_column_letter

class AggregationRequest(BaseModel):
    agg_columns: List[str]
    group_columns: List[str]
    operations: Dict[str, str]

app = FastAPI()

# Enable CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Mount static files
app.mount("/static", StaticFiles(directory="static"), name="static")

# Store uploaded files temporarily
UPLOAD_DIR = Path(tempfile.gettempdir()) / "excel_aggregator"
UPLOAD_DIR.mkdir(exist_ok=True)

# Mount the static directory
app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/")
async def read_root():
    return FileResponse("static/index.html")


@app.post("/process-table")
async def process_table(data: dict):
    df = pd.DataFrame(data['data'])
    return {"columns": list(df.columns)}


@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Only Excel files are supported")

    file_path = UPLOAD_DIR / file.filename

    # Handle pasted data
    if file.filename == 'pasted-data.xlsx':
        content = await file.read()
        data = loads(content.decode())
        df = pd.DataFrame(data['data'])
        
        # Save pasted data as Excel
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']
            # Format date column
            if 'Date' in df.columns:
                date_col = df.columns.get_loc('Date') + 1
                for cell in worksheet[get_column_letter(date_col)]:
                    cell.number_format = 'YYYY-MM-DD'
            # Format cost column
            if 'Cost' in df.columns:
                cost_col = df.columns.get_loc('Cost') + 1
                for cell in worksheet[get_column_letter(cost_col)]:
                    cell.number_format = '$#,##0.00'
    else:
        with open(file_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        df = pd.read_excel(file_path)

    columns = list(df.columns)
    return {"columns": columns, "filename": file.filename}


@app.post("/aggregate/{filename}")
async def aggregate_data(filename: str, request: AggregationRequest):
    file_path = UPLOAD_DIR / filename
    if not file_path.exists():
        raise HTTPException(404, "File not found")

    try:
        df = pd.read_excel(file_path)

        # Clean numeric columns before aggregation
        for col in request.agg_columns:
            if col in df.columns and isinstance(df[col].iloc[0], str):
                df[col] = pd.to_numeric(df[col].str.replace(
                    r'[^\d.]', '', regex=True), errors='coerce')

        result = df.groupby(request.group_columns)[
            request.agg_columns].agg(request.operations)

        output_filename = f"aggregated_{filename}"
        output_path = UPLOAD_DIR / output_filename
        result.to_excel(output_path)

        return FileResponse(
            output_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=output_filename
        )
    except Exception as e:
        raise HTTPException(500, str(e))
